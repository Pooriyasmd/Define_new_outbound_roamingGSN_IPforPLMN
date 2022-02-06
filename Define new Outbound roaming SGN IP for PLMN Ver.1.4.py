import openpyxl as xl
#the text file as MML you use should be named as "TEUGWH31.txt"
import ipaddress

def finding_SRV_NODE_PLMN():
    MMLfile = open("TEUGWH31.txt","r")
    ADDSRVPLMNFile = open("ADD SRVNODEPLMN.txt", "w")
    # read the whole text
    fulllineofsearch = MMLfile.readlines()
    lengthOfString = fulllineofsearch.__len__()
    SRVNODEPLMN = "ADD SRVNODEPLMN"
    index = 0
    for line in fulllineofsearch:
        index += 1
        if SRVNODEPLMN in line:
            ADDSRVPLMNFile.write(fulllineofsearch[index - 1])

    MMLfile.close()
    ADDSRVPLMNFile.close()


def Excel_file_creation():


    workbook = xl.load_workbook("ADD SRVNODEPLMN.xlsx")
    worksheet1 = workbook["Sheet1"]
    worksheet2 = workbook["Sheet2"]
    worksheet2.cell(1,1).value = "Start IP"
    worksheet2.cell(1,3).value = "End IP"
    worksheet2.cell(1,5).value = "PLMN"
    ADDSRVPLMNFile = open("ADD SRVNODEPLMN.txt", "r")
    ADDSRVPLMNread = ADDSRVPLMNFile.readlines()
    length = ADDSRVPLMNread.__len__()
    for i in range(1,length):
        worksheet1.cell(i, 1).value = ADDSRVPLMNread[i]
        valueforcellinworksheet2 = worksheet1.cell(i, 1).value
        valueforcellinworksheet2split = valueforcellinworksheet2.split('"')
    #   to copy each array of the list to excel file Sheet2
        for j in range(1, 6, 2):
            worksheet2.cell(i+1, j).value = valueforcellinworksheet2split[j]
    workbook.save("ADD SRVNODEPLMN.xlsx")
def createRangeIPfile():
    ADDSRVPLMNFile = open("ADD SRVNODEPLMN.txt", "r")
    ADDSRVPLMNread = ADDSRVPLMNFile.readlines()
    rangeIPsfile = open("RangeIP.txt", "w")
    rangeIPaslist = ["","  ","","  ","","\n"]
    for line in ADDSRVPLMNread:
        linesplit = line.split('"')
        rangeIPaslist[0] = linesplit[5]
        rangeIPaslist[2] = linesplit[1]
        rangeIPaslist[4] = linesplit[3]
        rangeIPsfile.writelines(rangeIPaslist)
    rangeIPsfile.close()

finding_SRV_NODE_PLMN()
#createRangeIPfile()


def ipcomparison(startIPofRangeIP, endIPofRangeIP, plmnofIPRange, startIPofIR21, endIPofIR21, plmnofIR21):
    if startIPofRangeIP == startIPofIR21 and endIPofIR21 == endIPofRangeIP:
        return f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnofIPRange}\n"
    if startIPofRangeIP == startIPofIR21+1 and endIPofIR21-1 == endIPofRangeIP:
        return f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnofIPRange}\n"
    if startIPofRangeIP == startIPofIR21 and endIPofIR21-1 == endIPofRangeIP:
        return f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnofIPRange}\n"
    if startIPofRangeIP == startIPofIR21+1 and endIPofIR21 == endIPofRangeIP:
        return f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnofIPRange}\n"
    if startIPofRangeIP <= startIPofIR21 <= endIPofRangeIP or startIPofRangeIP <= endIPofIR21 <= endIPofRangeIP:
        return f"{startIPofIR21} to {endIPofIR21} is PARTIALLY defined on plmn = {plmnofIPRange} as {startIPofRangeIP} to {endIPofRangeIP}\n"
    if startIPofRangeIP >= startIPofIR21 and endIPofRangeIP <= endIPofIR21:
        return f"{startIPofIR21} to {endIPofIR21} is PARTIALLY defined on plmn = {plmnofIPRange} as {startIPofRangeIP} to {endIPofRangeIP}\n"
    if startIPofRangeIP < startIPofIR21 and endIPofRangeIP < endIPofIR21:
        return f'ADD SRVNODEPLMN:IPVERSION=IPV4,SRVNODESTARTV4="{startIPofIR21}",SRVNODEENDV4="{endIPofIR21}",SRVNODEPLMN="{plmnofIR21}";\n'
    if startIPofRangeIP > startIPofIR21 and endIPofRangeIP > endIPofIR21:
        return f'ADD SRVNODEPLMN:IPVERSION=IPV4,SRVNODESTARTV4="{startIPofIR21}",SRVNODEENDV4="{endIPofIR21}",SRVNODEPLMN="{plmnofIR21}";\n'
    else:
        pass

def readIR21file(y):
    ir21file = open("IR21.txt","r")
    ir21 = ir21file.readlines()
    ir21filesplit = ir21[y].split("\t")
    plmn1 = ir21filesplit[2].split("\n")
    ip3 = ipaddress.IPv4Address(ir21filesplit[0])
    ip4 = ipaddress.IPv4Address(ir21filesplit[1])
    plmn = plmn1[0]
    return ip3, ip4, plmn

def readRangeIP(x):
    rangeIPsfile = open("RangeIP.txt", "r")
    rangeIPsfileread = rangeIPsfile.readlines()
    rangeIPsfilereadsplit = rangeIPsfileread[x].split(" ")
    plmn = rangeIPsfilereadsplit[0]
    ip1 = ipaddress.IPv4Address(rangeIPsfilereadsplit[2])
    ip12 = rangeIPsfilereadsplit[4].split("\n")
    ip2 = ipaddress.IPv4Address(ip12[0])
    return ip1, ip2, plmn

rangeIPsfile = open("RangeIP.txt", "r")
rangeIPsfileread = rangeIPsfile.readlines()
lengthofreadRangeIP = rangeIPsfileread.__len__()
ir21file = open("IR21.txt", "r")
ir21 = ir21file.readlines()
lengthofIR21file = ir21.__len__()
resultfile = open("Script.txt", "w")
for i in range(1, lengthofIR21file):
    resReadIR21 = readIR21file(i)
    startIPIR21 = ipaddress.IPv4Address(resReadIR21[0])
    endIPIR21 = ipaddress.IPv4Address(resReadIR21[1])
    plmnIR21 = resReadIR21[2]
    for j in range(0, lengthofreadRangeIP):
        resReadrange = readRangeIP(j)
        startIPRange = ipaddress.IPv4Address(resReadrange[0])
        endIPRange = ipaddress.IPv4Address(resReadrange[1])
        plmnIRrange = resReadrange[2]
        theresultofIPcomparison = ipcomparison(startIPRange, endIPRange, plmnIRrange,startIPIR21, endIPIR21, plmnIR21)
        resultfile.writelines(theresultofIPcomparison)
resultfile.close()
lines_seen = set() # holds lines already seen
outfile = open("ScriptFinal.txt", "w")
for line in open("Script.txt", "r"):
    if line not in lines_seen: # not a duplicate
        outfile.write(line)
        lines_seen.add(line)
outfile.close()








#finding_SRV_NODE_PLMN()
#Excel_file_creation()



