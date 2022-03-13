import openpyxl as xl
#the text file as MML you use should name it as "TEUGWH31.txt"
import ipaddress



def finding_SRV_NODE_PLMN():
    MMLfile = open("TEUGWH31.txt","r")
    ADDSRVPLMNFile = open("ADD SRVNODEPLMN.txt", "w")
    # read the whole text
    fulllineofsearch = MMLfile.readlines()
    SRVNODEPLMN = "ADD SRVNODEPLMN"
    index = 0
    for line in fulllineofsearch:
        index += 1
        if SRVNODEPLMN in line:
            ADDSRVPLMNFile.write(fulllineofsearch[index - 1])
    MMLfile.close()
    ADDSRVPLMNFile.close()


def createRangeIPfile():
    ADDSRVPLMNFile = open("ADD SRVNODEPLMN.txt", "r")
    ADDSRVPLMNread = ADDSRVPLMNFile.readlines()
    rangeIPsfile = open("RangeIP.txt", "w")
    rangeIPaslist = ["","  ","","  ","","\n"]
    for line in ADDSRVPLMNread:
        linesplit = line.split('"')
        rangeIPaslist[0] = linesplit[5]
        rangeIPaslist[2] = linesplit[1]
        rangeIPaslist[4] = linesplit[3]
        rangeIPsfile.writelines(rangeIPaslist)
    rangeIPsfile.close()

finding_SRV_NODE_PLMN()
createRangeIPfile()

def readRangeIP(x):
    rangeIPsfile = open("RangeIP.txt", "r")
    rangeIPsfileread = rangeIPsfile.readlines()
    rangeIPsfilereadsplit = rangeIPsfileread[x].split(" ")
    plmn = rangeIPsfilereadsplit[0]
    ip1 = ipaddress.IPv4Address(rangeIPsfilereadsplit[2])
    ip12 = rangeIPsfilereadsplit[4].split("\n")
    ip2 = ipaddress.IPv4Address(ip12[0])
    return ip1, ip2, plmn



def comparisonfunction(startIPofIR21, endIPofIR21, plmnofIR21):
    rangeIPsfile = open("RangeIP.txt", "r")
    rangeIPsfileread = rangeIPsfile.readlines()
    lengthofreadRangeIP = rangeIPsfileread.__len__()
    for i in range(0, lengthofreadRangeIP):
        resReadrange = readRangeIP(i)
        startIPRange = ipaddress.IPv4Address(resReadrange[0])
        endIPRange = ipaddress.IPv4Address(resReadrange[1])
        plmnIRrange = resReadrange[2]
        if startIPRange == startIPofIR21 and endIPofIR21 == endIPRange:
            checkresult = f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnIRrange}"
            break
        elif startIPRange == startIPofIR21 + 1 and endIPofIR21 - 1 == endIPRange:
            checkresult = f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnIRrange}"
            break
        elif startIPRange == startIPofIR21 and endIPofIR21 - 1 == endIPRange:
            checkresult = f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnIRrange}"
            break
        elif startIPRange == startIPofIR21 + 1 and endIPofIR21 == endIPRange:
            checkresult = f"{startIPofIR21} to {endIPofIR21} is defined on plmn = {plmnIRrange}"
            break
        elif startIPRange <= startIPofIR21 <= endIPRange or startIPRange <= endIPofIR21 <= endIPRange:
            checkresult = f"{startIPofIR21} to {endIPofIR21} is PARTIALLY defined on plmn = {plmnIRrange} as {startIPRange} to {endIPRange}"
            break
        elif startIPRange >= startIPofIR21 and endIPRange <= endIPofIR21:
            checkresult = f"{startIPofIR21} to {endIPofIR21} is PARTIALLY defined on plmn = {plmnIRrange} as {startIPRange} to {endIPRange}"
            break
        else:
            checkresult = f'ADD SRVNODEPLMN:IPVERSION=IPV4,SRVNODESTARTV4="{startIPofIR21}",SRVNODEENDV4="{endIPofIR21}",SRVNODEPLMN="{plmnofIR21}";'
    return checkresult


def supernetfunction(net1):
    for samerowcompare in range(4, worksheet.max_row + 1):
        ip = worksheet.cell(samerowcompare, 1).value
        net2 = ipaddress.IPv4Network(ip)
        if net1 == net2:
            continue
        if net1.subnet_of(net2):
            return f"Supernet found on {net2}"
        else:
            continue


workbook = xl.load_workbook("IR21.xlsx")
worksheet = workbook["IR21"]
worksheet.cell(3, 1).value = "IP Address Range"
worksheet.cell(1, 5).value = "By Pooriya Samadifar - P84145909"
worksheet.cell(3, 2).value = "TEUGWH31/TWUGWH31"
for row in range(4, worksheet.max_row + 1):
    ip = worksheet.cell(row, 1).value
    plmnofIR21 = worksheet.cell(2, 1).value
    numberofaddresses = ipaddress.IPv4Network(ip).num_addresses
    host = list(ipaddress.IPv4Network(ip).hosts())
    if numberofaddresses > 3:
        startIPofIR21 = host[0] - 1
        endIPofIR21 = host[numberofaddresses - 3] + 1
    elif numberofaddresses == 2:
        startIPofIR21 = host[0] - 1
        endIPofIR21 = host[1] + 1
    elif numberofaddresses == 1:
        startIPofIR21 = host[0]
        endIPofIR21 = host[0]
    worksheet.cell(row, 2).value = comparisonfunction(startIPofIR21, endIPofIR21, plmnofIR21)
for row in range(4, worksheet.max_row):
    ip = worksheet.cell(row, 1).value
    net1 = ipaddress.IPv4Network(ip)
    worksheet.cell(row, 3).value = supernetfunction(net1)
workbook.save("IR21.xlsx")






